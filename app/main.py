from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
import openpyxl
from typing import List
from app.grader.user_solution_tst import TestUserSolution, TestExtractor

app = FastAPI()

tasks_wb = openpyxl.load_workbook("app/data/tasks.xlsx")
open_hidden_wb = openpyxl.load_workbook("app/data/open_hidden.xlsx")

tasks_sheet = tasks_wb.active
open_tests_sheet = open_hidden_wb["Open"]
hidden_tests_sheet = open_hidden_wb["Hidden"]


class TaskResponse(BaseModel):
    task_id: int
    formulation: str
    example_code: str
    link_to_task: str

class LLMRecomendation(BaseModel):
    user_code: str
    errors: List[str]

class SolutionRequest(BaseModel):
    task_id: int
    solution_code: str 

class GetAllTasks(BaseModel):
    pass


@app.get("/task/{task_id}", response_model=TaskResponse)
def get_task(task_id: int):
    """
    Получает задачу с листа по номеру задачи.
    Возвращает условие задачи в виде str.
    """
    try:
        task_row = tasks_sheet[task_id + 1]  
        task = TaskResponse(
            task_id=task_id,
            formulation=task_row[1].value,
            example_code=task_row[4].value,
            link_to_task=task_row[3].value
        )
        return task
    except IndexError:
        raise HTTPException(status_code=404, detail="Task not found")

@app.get("/all_tasks/", response_model=List[str])
def get_all_tasks():
    """
    Возвращает номера всех задач в списке.
    """
    try:
        task_ids = [str(i + 1) for i, row in enumerate(tasks_sheet.iter_rows(min_row=1)) if row[0].value is not None]
        return task_ids
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal error: {str(e)}")



@app.post("/submit_solution/")
def submit_solution(request: SolutionRequest):
    """
    Прогоняет пользовательское решение по открытым и скрытым тестам.
    """
    try:
        solution_code = request.solution_code
        task_id = request.task_id
        open_test_extractor = TestExtractor(open_tests_sheet, task_id)
        hidden_test_extractor = TestExtractor(hidden_tests_sheet, task_id)

        open_tests = open_test_extractor.extract_tests()
        hidden_tests = hidden_test_extractor.extract_tests()

        test_runner = TestUserSolution(
                                test_cases=open_tests,
                                user_solution=solution_code
                                )

        open_test_errors, fail_open = test_runner.run_tests_with_unittest()

        if fail_open:
            return {
            "status": "failed",
            "message": f"Открытые тесты провалены {open_tests}. Ошибки: {'/n'.join(open_test_errors)}."
            }

        test_runner_hidden = TestUserSolution(
                                test_cases=hidden_tests,
                                user_solution=solution_code
                                )
    
        _, fail_hidden = test_runner_hidden.run_tests_with_unittest()

        if fail_hidden:
            return {
            "status": "failed",
            "message": "Скрытые тесты провалены."
            }

        reference_solution = tasks_sheet[task_id + 1][2].value.replace("\\n", "\n")

        if solution_code != reference_solution:

            return {
            "status": "passed",
            "user_code": f"{solution_code}",
            "ideal_code": f"{reference_solution}",
            "message": f"Тесты прошли успешно, а решения не совпадают."
            }

        return {
        "status": "success",
        "message": "Все тесты пройдены успешно!"
        }
        
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal error: {str(e)}")
